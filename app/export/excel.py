"""Exportación a Excel (openpyxl) de todos los entregables de una empresa.

Genera un libro con una hoja por módulo, conservando la estructura de los
Excel originales: Resumen, Matrices, Planes, KPI y CMI.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.core import semaforo
from app.models.cmi import Perspectiva
from app.models.empresa import Empresa
from app.models.matriz import Matriz
from app.services.dashboard import _estado_indicador, generar_dashboard_general
from app.services.planes import totalizar_consolidado_empresa

_FILL_ENCABEZADO = PatternFill("solid", fgColor="1F4E78")
_FONT_ENCABEZADO = Font(bold=True, color="FFFFFF")
_FONT_TITULO = Font(bold=True, size=13, color="1F4E78")
_ALINEACION_ENCABEZADO = Alignment(horizontal="center", vertical="center")


def generar_excel_empresa(db: Session, empresa_id: int) -> BytesIO | None:
    """Devuelve un buffer xlsx con los entregables de la empresa, o None si no existe."""
    empresa = db.get(Empresa, empresa_id)
    if empresa is None:
        return None

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja por defecto

    _hoja_resumen(wb, db, empresa)
    _hoja_matrices(wb, db, empresa_id)
    _hoja_planes(wb, db, empresa_id)
    _hoja_kpi(wb, db, empresa_id)
    _hoja_cmi(wb, db, empresa_id)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _escribir_encabezado(ws: Worksheet, fila: int, columnas: list[str]) -> None:
    for col, titulo in enumerate(columnas, start=1):
        celda = ws.cell(row=fila, column=col, value=titulo)
        celda.fill = _FILL_ENCABEZADO
        celda.font = _FONT_ENCABEZADO
        celda.alignment = _ALINEACION_ENCABEZADO


def _autoajustar(ws: Worksheet, anchos: list[int]) -> None:
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho


def _hoja_resumen(wb: Workbook, db: Session, empresa: Empresa) -> None:
    ws = wb.create_sheet("Resumen")
    ws["A1"] = f"Resumen ejecutivo — {empresa.nombre}"
    ws["A1"].font = _FONT_TITULO

    data = generar_dashboard_general(db, empresa.id)
    filas = [
        ("Matrices", "Total de matrices", data.matrices.total),
        ("Planes", "Planes", data.planes.total_planes),
        ("Planes", "Estrategias", data.planes.total_estrategias),
        ("Planes", "Actividades", data.planes.total_actividades),
        ("Planes", "Costo total", data.planes.total_costo),
        ("KPI", "Indicadores tácticos", data.kpi.total_indicadores),
        ("CMI", "Perspectivas", data.cmi.total_perspectivas),
        ("CMI", "Objetivos", data.cmi.total_objetivos),
        ("CMI", "Indicadores", data.cmi.total_indicadores),
        ("CMI", "Mediciones", data.cmi.total_mediciones),
        ("CMI", "Semáforos en verde", data.cmi.semaforos.verde),
        ("CMI", "Semáforos en amarillo", data.cmi.semaforos.amarillo),
        ("CMI", "Semáforos en rojo", data.cmi.semaforos.rojo),
        ("CMI", "Sin medición", data.cmi.semaforos.sin_medicion),
        ("Trazabilidad", "Estrategias con matriz origen", data.trazabilidad.estrategias_con_matriz),
        ("Trazabilidad", "Indicadores CMI con KPI", data.trazabilidad.indicadores_cmi_con_kpi),
    ]
    _escribir_encabezado(ws, 3, ["Módulo", "Métrica", "Valor"])
    for i, (modulo, metrica, valor) in enumerate(filas, start=4):
        ws.cell(row=i, column=1, value=modulo)
        ws.cell(row=i, column=2, value=metrica)
        ws.cell(row=i, column=3, value=valor)
    _autoajustar(ws, [16, 34, 14])


def _hoja_matrices(wb: Workbook, db: Session, empresa_id: int) -> None:
    ws = wb.create_sheet("Matrices")
    _escribir_encabezado(
        ws, 1, ["Matriz", "Tipo", "Factor", "Peso", "Calificación", "Resultado"]
    )
    fila = 2
    matrices = db.query(Matriz).filter(Matriz.empresa_id == empresa_id).all()
    for matriz in matrices:
        if not matriz.factores:
            ws.cell(row=fila, column=1, value=matriz.nombre)
            ws.cell(row=fila, column=2, value=matriz.tipo.value)
            fila += 1
            continue
        for factor in matriz.factores:
            ws.cell(row=fila, column=1, value=matriz.nombre)
            ws.cell(row=fila, column=2, value=matriz.tipo.value)
            ws.cell(row=fila, column=3, value=factor.descripcion)
            ws.cell(row=fila, column=4, value=factor.peso)
            ws.cell(row=fila, column=5, value=factor.calificacion)
            ws.cell(row=fila, column=6, value=factor.resultado)
            fila += 1
    _autoajustar(ws, [26, 14, 40, 10, 14, 12])


def _hoja_planes(wb: Workbook, db: Session, empresa_id: int) -> None:
    ws = wb.create_sheet("Planes")
    _escribir_encabezado(
        ws,
        1,
        ["Plan", "Tipo de estrategia", "Estrategia", "Actividad", "Responsable",
         "Tiempo", "Costo", "Tipo de cuenta"],
    )
    fila = 2
    consolidado = totalizar_consolidado_empresa(db, empresa_id)
    from app.models.plan import Plan  # import local para evitar ciclos innecesarios

    planes = db.query(Plan).filter(Plan.empresa_id == empresa_id).all()
    for plan in planes:
        for estrategia in plan.estrategias:
            if not estrategia.actividades:
                ws.cell(row=fila, column=1, value=plan.tipo.value)
                ws.cell(row=fila, column=2, value=estrategia.tipo_estrategia)
                ws.cell(row=fila, column=3, value=estrategia.descripcion)
                fila += 1
                continue
            for actividad in estrategia.actividades:
                ws.cell(row=fila, column=1, value=plan.tipo.value)
                ws.cell(row=fila, column=2, value=estrategia.tipo_estrategia)
                ws.cell(row=fila, column=3, value=estrategia.descripcion)
                ws.cell(row=fila, column=4, value=actividad.descripcion)
                ws.cell(row=fila, column=5, value=actividad.responsable)
                ws.cell(row=fila, column=6, value=actividad.tiempo)
                ws.cell(row=fila, column=7, value=actividad.costo)
                ws.cell(row=fila, column=8, value=actividad.tipo_cuenta)
                fila += 1

    fila += 1
    celda_total = ws.cell(row=fila, column=6, value="TOTAL")
    celda_total.font = Font(bold=True)
    valor_total = ws.cell(row=fila, column=7, value=consolidado.total_costo)
    valor_total.font = Font(bold=True)
    _autoajustar(ws, [14, 20, 36, 32, 18, 14, 12, 16])


def _hoja_kpi(wb: Workbook, db: Session, empresa_id: int) -> None:
    ws = wb.create_sheet("KPI")
    _escribir_encabezado(
        ws, 1, ["Estrategia", "Indicador", "Fórmula", "Frecuencia", "Ponderación"]
    )
    fila = 2
    from app.models.plan import Estrategia, Indicador, Plan

    indicadores = (
        db.query(Indicador, Estrategia)
        .join(Estrategia, Indicador.estrategia_id == Estrategia.id)
        .join(Plan, Estrategia.plan_id == Plan.id)
        .filter(Plan.empresa_id == empresa_id)
        .all()
    )
    for indicador, estrategia in indicadores:
        ws.cell(row=fila, column=1, value=estrategia.descripcion)
        ws.cell(row=fila, column=2, value=indicador.nombre)
        ws.cell(row=fila, column=3, value=indicador.formula)
        ws.cell(row=fila, column=4, value=indicador.frecuencia)
        ws.cell(row=fila, column=5, value=indicador.ponderacion)
        fila += 1
    _autoajustar(ws, [36, 30, 40, 14, 14])


def _hoja_cmi(wb: Workbook, db: Session, empresa_id: int) -> None:
    ws = wb.create_sheet("CMI")
    _escribir_encabezado(
        ws,
        1,
        ["Perspectiva", "Objetivo", "Indicador", "Meta", "Sentido",
         "Última medición", "Semáforo"],
    )
    fila = 2
    perspectivas = (
        db.query(Perspectiva).filter(Perspectiva.empresa_id == empresa_id).all()
    )
    for perspectiva in perspectivas:
        for objetivo in perspectiva.objetivos:
            for indicador in objetivo.indicadores:
                ultima = (
                    max(indicador.mediciones, key=lambda m: m.id).valor
                    if indicador.mediciones
                    else None
                )
                estado = _estado_indicador(indicador)
                ws.cell(row=fila, column=1, value=perspectiva.tipo.value)
                ws.cell(row=fila, column=2, value=objetivo.descripcion)
                ws.cell(row=fila, column=3, value=indicador.nombre)
                ws.cell(row=fila, column=4, value=indicador.meta)
                ws.cell(row=fila, column=5, value=indicador.sentido.value)
                ws.cell(row=fila, column=6, value=ultima)
                celda_estado = ws.cell(
                    row=fila, column=7,
                    value=estado.value if estado is not None else "SIN MEDICIÓN",
                )
                _pintar_semaforo(celda_estado, estado)
                fila += 1
    _autoajustar(ws, [16, 34, 30, 10, 10, 16, 16])


def _pintar_semaforo(celda, estado: semaforo.Estado | None) -> None:
    colores = {
        semaforo.Estado.VERDE: "2E9E4F",
        semaforo.Estado.AMARILLO: "E8B53A",
        semaforo.Estado.ROJO: "D1493F",
    }
    color = colores.get(estado)
    if color is not None:
        celda.fill = PatternFill("solid", fgColor=color)
        celda.font = Font(bold=True, color="FFFFFF")
