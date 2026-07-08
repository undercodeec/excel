"""Tests de exportación a Excel (Fase 6)."""
from io import BytesIO

from openpyxl import load_workbook


def _crear_empresa_directa(client):
    from app.database import get_db
    from app.models.empresa import Empresa

    db = next(client.app.dependency_overrides[get_db]())
    emp = Empresa(nombre="Hacienda Celia Maria C.A.")
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp.id


def _crear_matriz_directa(client, empresa_id):
    from app.database import get_db
    from app.models.enums import TipoMatriz
    from app.models.matriz import Matriz, FactorMatriz

    db = next(client.app.dependency_overrides[get_db]())
    matriz = Matriz(empresa_id=empresa_id, tipo=TipoMatriz.efi, nombre="EFI")
    matriz.factores.append(
        FactorMatriz(descripcion="Liquidez", peso=1.0, calificacion=3, resultado=3.0)
    )
    db.add(matriz)
    db.commit()
    db.refresh(matriz)
    return matriz.id


def test_export_excel_contiene_todas_las_hojas(client):
    emp_id = _crear_empresa_directa(client)
    _crear_matriz_directa(client, emp_id)

    client.post(
        "/api/planes",
        json={
            "empresa_id": emp_id,
            "tipo": "marketing",
            "estrategias": [
                {
                    "descripcion": "Abrir canales",
                    "actividades": [{"descripcion": "Mapear", "costo": 1500, "responsable": "Comercial"}],
                    "indicadores": [{"nombre": "Mercados abiertos", "formula": "(a/b)", "frecuencia": "mensual", "ponderacion": 1.0}],
                }
            ],
        },
    )
    client.post(
        "/api/cmi/perspectivas",
        json={
            "empresa_id": emp_id,
            "tipo": "financiera",
            "nombre": "Financiera",
            "objetivos": [
                {
                    "descripcion": "Rentabilidad",
                    "indicadores": [
                        {
                            "nombre": "Margen",
                            "meta": 0.2,
                            "rango_min": 0.1,
                            "rango_max": 0.2,
                            "mediciones": [{"periodo": "2026-07", "valor": 0.25}],
                        }
                    ],
                }
            ],
        },
    )

    r = client.get(f"/api/export/excel?empresa_id={emp_id}")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in r.headers["content-disposition"]
    assert f"empresa_{emp_id}.xlsx" in r.headers["content-disposition"]

    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Resumen", "Matrices", "Planes", "KPI", "CMI"]

    # Matrices: factor exportado
    matrices = wb["Matrices"]
    assert matrices["A2"].value == "EFI"
    assert matrices["C2"].value == "Liquidez"

    # Planes: actividad + fila TOTAL
    planes = wb["Planes"]
    valores_planes = [c.value for row in planes.iter_rows() for c in row]
    assert "Mapear" in valores_planes
    assert "TOTAL" in valores_planes
    assert 1500 in valores_planes

    # KPI
    kpi = wb["KPI"]
    assert kpi["B2"].value == "Mercados abiertos"

    # CMI: semáforo verde (0.25 >= meta 0.2)
    cmi = wb["CMI"]
    assert cmi["C2"].value == "Margen"
    assert cmi["G2"].value == "VERDE"


def test_export_excel_empresa_vacia(client):
    emp_id = _crear_empresa_directa(client)
    r = client.get(f"/api/export/excel?empresa_id={emp_id}")
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames == ["Resumen", "Matrices", "Planes", "KPI", "CMI"]


def test_export_excel_empresa_inexistente(client):
    assert client.get("/api/export/excel?empresa_id=999").status_code == 404
